from pathlib import Path
from django.core.exceptions import ValidationError
from django.core.management import BaseCommand, CommandError
from django.core.validators import validate_email
from openpyxl import load_workbook

from parsers.models import Letters


class Command(BaseCommand):
    """Import xlsx file and run email sender."""

    required_columns = ("external_id", "user_id", "email", "subject", "message")

    batch_size = 5

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to XLSX file.")

    def handle(self, *args, **options):
        """
        Runs the full import script.

        - Checks the input file;
        - reads XLSX rows in a stream;
        - validates and batches the data;
        - saves the new records;
        - displays the final statistics.
        """
        file_path = Path(options["file_path"])
        self.validate_options(file_path)

        stats = {
            "processed_rows": 0,
            "created_records": 0,
            "skipped_records": 0,
            "error_rows": 0,
        }
        batch: list[dict[str, str]] = []
        seen_external_ids: set[str] = set()

        workbook = None
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)

            headers = next(rows, None)
            if not headers:
                raise CommandError("The XLSX file is empty.")

            column_indexes = self.normalize_column_indexes(headers)

            for row_number, row in enumerate(rows, start=2):
                stats["processed_rows"] += 1

                validated_data = self.validate_row(
                    row=row,
                    row_number=row_number,
                    column_indexes=column_indexes,
                )
                if validated_data is None:
                    stats["error_rows"] += 1
                    continue

                external_id = validated_data["external_id"]
                if external_id in seen_external_ids:
                    stats["skipped_records"] += 1
                    continue

                seen_external_ids.add(external_id)
                batch.append(validated_data)

                if len(batch) >= self.batch_size:
                    self.push_batch_data(batch=batch, stats=stats)

            if batch:
                self.push_batch_data(batch=batch, stats=stats)
        except CommandError:
            raise
        except Exception as exc:
            raise CommandError(f"Error while processing XLSX file: {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()

        self.stdout.write(self.style.SUCCESS("Import finished."))
        self.stdout.write(f"Processed rows: {stats['processed_rows']}")
        self.stdout.write(f"Created records: {stats['created_records']}")
        self.stdout.write(f"Skipped records: {stats['skipped_records']}")
        self.stdout.write(f"Error rows: {stats['error_rows']}")

    def validate_options(self, file_path: Path) -> None:
        """Checks that the file exists and the usual file"""
        if not file_path.exists() or not file_path.is_file():
            raise CommandError(f"File not found: {file_path}")

    def normalize_column_indexes(self, headers: tuple) -> dict:
        """
        Creates a map of column indices based on the column headers.

        Throws a `CommandError` if at least one required column is missing.
        """
        header_map = {
            str(value).strip().lower(): index
            for index, value in enumerate(headers)
            if value is not None and str(value).strip()
        }
        missing_columns = [column for column in self.required_columns if column not in header_map]
        if missing_columns:
            raise CommandError(f"Missing required columns: {missing_columns}")
        return {column: header_map[column] for column in self.required_columns}

    def validate_row(self, row: tuple, row_number: int, column_indexes: dict) -> dict | None:
        """
        Normalizes and validates a single line of a file.

        Returns a dictionary containing the data if the line is valid.
        Returns `None` if there is an error (and writes the reason to stderr).
        """
        data: dict = {}
        for column in self.required_columns:
            idx = column_indexes[column]
            value = row[idx] if idx < len(row) else None
            data[column] = "" if value is None else str(value).strip()

        if not all(data.values()):
            self.stderr.write(f"Row {row_number}: missing one or more required fields.")
            return None

        try:
            validate_email(data["email"])
        except ValidationError:
            self.stderr.write(f"Row {row_number}: invalid email '{data['email']}'.")
            return None

        return data

    def push_batch_data(self, batch: list, stats: dict) -> None:
        """
         Processes the accumulated batch of records.

        - Excludes records that already exist in the database by `external_id`;
        - creates new records using `bulk_create`;
        - initiates the submission of new records;
        - Updates the statistics and clears the batch.
        """
        external_ids = [row["external_id"] for row in batch]
        existing_ids = set(
            Letters.objects.filter(external_id__in=external_ids).values_list(
                "external_id", flat=True
            )
        )

        new_objects: list[Letters] = []
        for row in batch:
            if row["external_id"] in existing_ids:
                stats["skipped_records"] += 1
                continue

            new_objects.append(
                Letters(
                    external_id=row["external_id"],
                    user_id=row["user_id"],
                    email=row["email"],
                    subject=row["subject"],
                    message=row["message"],
                )
            )

        if new_objects:
            Letters.objects.bulk_create(new_objects, batch_size=self.batch_size)
            stats["created_records"] += len(new_objects)

        batch.clear()
